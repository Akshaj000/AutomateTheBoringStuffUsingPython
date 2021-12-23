#!python3
#Spreadsheetcellinverter.py - inverts column and row 
#insert M blank rows into the spreadsheet starting at row N

import openpyxl,sys
if len(sys.argv) > 1:
    filename = sys.argv[1]
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    tuple = tuple(sheet)
    count = 0
    for i in tuple:
        for j in i:
            c = j.column
            r = j.row
            sheet2.cell(row=c, column =r).value = sheet.cell(row=r, column=c).value 
            
    wb2.save(filename+'2.xlsx')
        
    
else:
    print("Enter arguement n!")