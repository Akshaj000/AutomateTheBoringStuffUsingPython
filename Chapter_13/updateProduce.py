#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

Price_Updates = {'Garlic':3.07,'Celery':1.19,'Lemon':1.27}

for rowno in range(2,sheet.max_row):
    produceName = sheet.cell(row=rowno,column=1).value
    if produceName in Price_Updates:
        sheet.cell(row=rowno,column=2).value=Price_Updates[produceName]
wb.save('updatedProduceSales.xlsx')